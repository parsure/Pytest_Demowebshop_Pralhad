import openpyxl

class XLUtils:

    @staticmethod
    def getRowCount(file, sheetName):
        workbook = openpyxl.load_workbook(file)
        sheet = workbook[sheetName]
        rowCount = sheet.max_row
        return rowCount
    @staticmethod
    def getColumnCount(file, shhetName):
        workbook = openpyxl.load_workbook(file)
        sheet = workbook[shhetName]
        rowCount = sheet.max_column
        return rowCount
    @staticmethod
    def readData(file, sheetName, rowNo, columnNo):
        workbook = openpyxl.load_workbook(file)
        sheet = workbook[sheetName]
        return sheet.cell(row=rowNo, column=columnNo).value

